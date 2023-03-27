# Import all the modules; import all the functions

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import pyexcel
import os
from time import sleep

import yaml

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList



def count_rows(ws):
    """
    Parameters
    ----------
    ws : worksheet object
    
    Returns
    -------
    number of populated rows in worksheet. 

    """
    count_row = 0

    for row in ws:
        if not all([cell.value is None for cell in row]):
            count_row += 1

    return(count_row)


def count_cols(ws, char=True):
    """ to be written """
    
    count_column = 0
    for column in ws.iter_cols():
        if not all([cell.value is None for cell in column]):
            count_column += 1
    
    
    if char:
        return get_column_letter(count_column)
    else:
        return count_column
    
    
def scrape_table(ws, content_range):
    """
    Parameters
    ----------
    ws: current worksheet
    content_range: the range of cells you want in the list.
    
    returns
    ----------
    list with all the info from the selected content range.

    """
    scraped_info = []
    
    for row in ws[content_range]:
        for cell in row:
            scraped_info.append(cell.value)

    return scraped_info



def write_table(ws, rows, row_range, cols = None, col_range = None):
    """
    Parameters:
    ----------
    ws: current worksheet
    rows: list with data for the rows
    cols: list with the data for the columns. (optional)
    
    row_range: range for the "rows" content output.
    col_range: range for the "cols" content output.(optional)
    ----------
    Returns:
    ----------
    Nothing. But writes the data on the excel sheet. 

    If you just want to output a list as a table, dont use the col params. 
    If you want to build the headers on top and at the side of the table, 
    use all the parameters.
    """
    # write the tables.
    j = 0
    for i, row in enumerate(ws[row_range]):
        for cell in row:
            cell.value = rows[j]
            j += 1

    
    if cols == None:
        pass
    
    else:
        try:
            j = 0
            for col in ws[col_range]:
                
                for cell in col:
                    
                    cell.value = cols[j]
                    j += 1
        
        except:
            pass            
              
        
def erase_table(ws, row_range):
    """
    to be written
    """
    # write the tables.
    j = 0
    for i, row in enumerate(ws[row_range]):
        for cell in row:
            cell.value = none
            j += 1

    
   


        
        
def mod_font(ws, cell_range, fsize=15, fbold=True):
    """
    to be written
    """
    
    for i, row in enumerate(ws[cell_range]):
        for cell in row:
            cell.font = Font(size=fsize, bold=fbold)
          
            
          
            
           
def mod_color(ws, cell_range, color='DAEEF3'):
    
    for i, row in enumerate(ws[cell_range]):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid',
                                        start_color=color,
                                        end_color=color)
          
                  

def set_border(ws, cell_range, _thin=True, _medium=False, _thick=False):
    
    thick = Side(border_style="thick", color="000000")
    medium = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    
    
    for row in ws[cell_range]:
        
        if _thick == True:
            for cell in row:
                cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
                
        elif _medium == True:
            for cell in row:
                cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)
        
        elif _thin == True:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        else:
            pass


def convert_xls(xls_file):
    """
    xls_file : xls file.
    
    ------------
        takes an XLS file. 
    ------------
        creates a converted XLSX file.
    
    """
    if "xlsx" in xls_file:
        pass
    else:
        pyexcel.save_book_as(file_name=xls_file,
                             dest_file_name=xls_file+"x")
        
    return xls_file + "x"


def delete_file(file):
    """
    file : any file.
    ----------
        takes a file.
    ----------
        deletes it. 
    
    """

    os.remove(file)




def date_conversion(date_string):
    """
    Parameters
    ----------
    date_string : String containing a date. example:
        "110223"  or  "11/02/21"  or  "blahblahbvlah 110423"

    Returns
    -------
    DateTime Object.
    """
    # strip the symbols in between the dates if existant.
    
    date = date_string.split("-")
    
    
    y = int(date[2].split(".")[0])
    m = int(date[1])
    d = int(date[0])

    return datetime(y, m, d)



def date_conversion_rev(dt):
    """
    Parameters
    -------
    datetime: datetime object.

    Returns
    -------
    String containing a datestring in the following format: "YYMMDD"
    """
    
    return dt.strftime("%m%d%y")
    
    
def move_sheet(wb, from_loc=None, to_loc=None):
    sheets=wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    #if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)
    sleep(0.5)
    
    
def get_file_names():
    """
    Returns
    -------
    All filenames of this python file's current directory, excluding itself. 
    
    """

    current_dir = os.getcwd()

    file_list = os.listdir(current_dir)

    return file_list


def get_last_week(datetime):
    """
    Parameters
    ----------
    datetime : datetime object. 

    Returns
    -------
    datetime_lastweek : datetime object containing last week's date in
    relation to the received datetime.
    """
    return (datetime - timedelta(days=7))



def get_week_id(dt):
    return dt.strftime("%V")

def get_week_str(dt):
    return dt.strftime("%m/%d/%y")



